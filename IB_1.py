### STEP0. 라이브러리 가져오기
import requests
import zipfile
import os
import pandas as pd
from bs4 import BeautifulSoup
from io import BytesIO
import pickle
import warnings
import json
import re
import streamlit as st
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, colors, Color, Font
from streamlit_lottie import st_lottie
from time import sleep

warnings.filterwarnings(action='ignore')
API_KEY = '9ed5bc3d6d1fee4f927c5d6a44eb5368d222824c'

### STEP1. 증권신고서(체무증권)의 최초신고서와 정정신고서 정보 가져오기
def get_info(bgn_de, end_de) :
    rcept_info = []
    url = 'https://opendart.fss.or.kr/api/list.xml'
    params = {'crtfc_key': API_KEY
              , 'bgn_de' : bgn_de
              , 'end_de' : end_de
              , 'pblntf_detail_ty': 'C002' # 증권신고(채무증권)
              , 'last_reprt_at':'N'} # 최종보고서 검색여부
    response = requests.get(url, params=params, verify=False)
    soup = BeautifulSoup(response.content, features='xml')
    total_page = soup.find('total_page').get_text()
    
    for i in range(1, int(total_page) + 1):
        params = {'crtfc_key': API_KEY
                  , 'bgn_de': bgn_de
                  , 'end_de': end_de
                  , 'page_no': str(i)
                  , 'pblntf_detail_ty': 'C002' # 증권신고(채무증권)
                  , 'last_reprt_at':'N'} # 최종보고서 검색여부
        response = requests.get(url, params=params, verify=False)
        soup = BeautifulSoup(response.content, features='xml')
        rcept_names = ['증권신고서(채무증권)','[첨부추가]증권신고서(채무증권)','[첨부정정]증권신고서(채무증권)','[기재정정]증권신고서(채무증권)']
        for c in soup.find_all('list'):
            if c.report_nm.get_text() in rcept_names:
                if c.report_nm.get_text()=='[기재정정]증권신고서(채무증권)':
                    rcept_info.append(c.rcept_no.get_text()+'_'+c.corp_code.get_text()+'_'+c.corp_name.get_text()+'_정정')
                else : 
                    rcept_info.append(c.rcept_no.get_text()+'_'+c.corp_code.get_text()+'_'+c.corp_name.get_text()+'_최초')
    print('보고서 수 : ', len(rcept_info))  
    st.write('보고서 수 : ', len(rcept_info)) 
    return(rcept_info)

### STEP2. STEP1에서 가져온 정정신고서의 최초신고서 가져오기
def get_rcept_no(info) :
    rcept_all = []
    url = 'https://opendart.fss.or.kr/api/list.xml'
    params = {'crtfc_key': API_KEY
              , 'bgn_de' : str(datetime.strptime(info[:8],'%Y%m%d') - timedelta(days=50)).replace('-','')[:8]
              , 'end_de' : info[:8]
              , 'corp_code' : info[15:23]
              , 'pblntf_detail_ty': 'C002' #증권신고(채무증권)
              , 'last_reprt_at':'N'} # 최종보고서 검색여부
    response = requests.get(url, params=params, verify=False)
    soup = BeautifulSoup(response.content, features='xml')
    total_page = soup.find('total_page').get_text()
    
    for i in range(1, int(total_page) + 1):
        params = {'crtfc_key': API_KEY
                  , 'bgn_de': str(datetime.strptime(info[:8],'%Y%m%d') - timedelta(days=50)).replace('-','')[:8]
                  , 'end_de': info[:8]
                  , 'page_no': str(i)
                  , 'corp_code' : info[15:23]
                  , 'pblntf_detail_ty': 'C002' #증권신고(채무증권)
                  , 'last_reprt_at':'N'} # 최종보고서 검색여부
        response = requests.get(url, params=params, verify=False)
        soup = BeautifulSoup(response.content, features='xml')
        rcept_names = ['증권신고서(채무증권)','[첨부추가]증권신고서(채무증권)','[첨부정정]증권신고서(채무증권)']
        for c in soup.find_all('list'):
            if c.report_nm.get_text() in rcept_names: 
                rcept_all.append(c.rcept_no.get_text())
    return(rcept_all[0])

### STEP3. 신고서 데이터 수집하기
def get_corp_docu(rcept_no):
    url = 'https://opendart.fss.or.kr/api/document.xml'
    params = {'crtfc_key': API_KEY, 'rcept_no': rcept_no}
    response = requests.get(url, params=params)
       
    try:
        zf = zipfile.ZipFile(BytesIO(response.content))
        z_list = zf.namelist()
        file = zf.read(z_list[0]) 
        
        soup = BeautifulSoup(file, 'html.parser', from_encoding='utf-8')
        table = soup.find('table-group', attrs={'aclass':'PL_KND_WRT'})
        company_nm = soup.find('company-name').get_text() 
        rows = [] 
        knd_wrt = table.find('tu', attrs={'aunit':'PL_KND_WRT'}).get_text()
        num = len(soup.find_all('table-group', attrs={'aclass':'PL_KND_WRT'}))
        for i in reversed(range(num)):
            table1 = soup.find_all('table-group', attrs={'aclass':'PL_KND_WRT'})[i]
            table2 = soup.find_all('table-group', attrs={'aclass':'SCHD'})[i]
            table3 = soup.find_all('table-group', attrs={'aclass':'ACC'})[i]

            seq_no = table1.find('te', attrs={'acode':'SEQ_NO'}).get_text()
            if rcept_no[:8] >= '20230220': #2023년 2월 20일 이후부터 신용등급 산출양식이 바뀜
                num = len(soup.find_all('table-group', attrs={'aclass':'CR_HIS'})[0].find_all('tu', attrs={'aunit':'CR_GRD'}))
                crdt_grd = []
                for j in range(num):
                    table4 = soup.find_all('table-group', attrs={'aclass':'CR_HIS'})[i]
                    crdt_grd.append(table4.find_all('tu', attrs={'aunit':'CR_GRD'})[j].get_text())
            else:
                crdt_grd = table1.find('te', attrs={'acode':'CRDT_GRD'}).get_text()           
            exp_dt = table1.find('tu', attrs={'aunit':'EXP_DT'}).get_text()      
            pym_prd = table2.find('tu', attrs={'aunit':'PYM_PRD'}).get_text()                     
            face_tot = table1.find('te', attrs={'acode':'FACE_TOT'}).get_text().strip().replace(',','')[:-8]          
            num = len(table3.find_all('tu', attrs={'aunit':'ACC_KND'}))
            acc_nmt1 = []
            acc_cnt1 = []
            acc_nmt2 = []
            acc_cnt2 = []
            for j in range(num):
                if '대표' in table3.find_all('tu', attrs={'aunit':'ACC_KND'})[j].get_text() or '공동' in table3.find_all('tu', attrs={'aunit':'ACC_KND'})[j].get_text():
                    acc_nmt1.append(table3.find_all('tu', attrs={'aunit':'ACC_NMT'})[j].get_text())
                    acc_cnt1.append(table3.find_all('te', attrs={'acode':'ACC_CNT'})[j].get_text().strip().replace(',','')[:-4])
                if '인수' in table3.find_all('tu', attrs={'aunit':'ACC_KND'})[j].get_text():
                    acc_nmt2.append(table3.find_all('tu', attrs={'aunit':'ACC_NMT'})[j].get_text())
                    acc_cnt2.append(table3.find_all('te', attrs={'acode':'ACC_CNT'})[j].get_text().strip().replace(',','')[:-4])

            row = {'채무증권명칭':knd_wrt,'신고서':rcept_no[:8],'발행사':company_nm,'회차':seq_no,'신용등급':crdt_grd,
                   '상환기일':exp_dt,'납입기일':pym_prd,'수량':face_tot,'대표':acc_nmt1, '대표수량':acc_cnt1,'인수':acc_nmt2,'인수수량':acc_cnt2}
            rows.append(row)    
    except Exception as e:
        print(rcept_no+'_Error!_'+str(e))   
    return(rows)

### STEP4. 원하는 형식으로 정리하기
def get_report(info):
    rows=[]
    for i in range(len(info)):
        try:
            if info[i][-2:]=='정정':            
                first = get_corp_docu(get_rcept_no(info[i]))
                correct = get_corp_docu(info[i][:14])
            else :
                first = get_corp_docu(info[i][:14])
                correct = []
            for j in range(len(first)):
                company_nm = first[j]['발행사'].replace('(주)','').replace('㈜','').replace(' 주식회사','').replace('주식회사 ','').replace('주식회사','')
                seq_no = first[j]['회차'].replace(' ','')
                if type(first[j]['신용등급']) == list:
                    crdt_grd = '/'.join(set(list(re.compile('\(([^)]+)').findall(str(first[0]['신용등급'])))))
                else :
                    crdt_grd = first[j]['신용등급'].replace('&cr','')
                due =  '-' if first[j]['상환기일'] == '-' else round((datetime.strptime(re.sub(r'[^0-9]', '', first[j]['상환기일']),'%Y%m%d')-datetime.strptime(re.sub(r'[^0-9]', '', first[j]['납입기일']),'%Y%m%d')).days/365,1)
                rcept_dt = datetime.strptime(first[j]['신고서'],'%Y%m%d')
                pym_prd = datetime.strptime(re.sub(r'[^0-9]', '', first[j]['납입기일']),'%Y%m%d')   
                face_tot1 = int(first[j]['수량'])
                face_tot2 = '' if correct == [] else int(correct[j]['수량'])
                acc_nmt1 = '\n'.join(first[j]['대표']).replace('투자증권','').replace('금융투자','').replace('에셋증권','').replace('증권','').replace('아이비케이','IBK').replace('케이비','KB').replace('한국산업은행','산은')
                acc_nmt2 = '\n'.join(first[j]['인수']).replace('투자증권','').replace('금융투자','').replace('에셋증권','').replace('증권','').replace('아이비케이','IBK').replace('케이비','KB').replace('한국산업은행','산은')
                    
                if correct != [] :
                    if first[j]['회차'].replace(' ','') != correct[j]['회차'].replace(' ',''):
                        acc_cnt1 = '회차불일치'
                        acc_cnt2 = '회차불일치'
                    else:
                        if first[j]['대표'] == correct[j]['대표']:
                            acc_cnt1 = '\n'.join([aj if (aj == bj) else str(aj)+'(*)' for aj, bj in zip(correct[j]['대표수량'], first[j]['대표수량'])])
                        else :
                            acc_cnt1 = '확인필요' 
                        if first[j]['인수'] == correct[j]['인수']:            
                            acc_cnt2 = '\n'.join([aj if (aj == bj) else str(aj)+'(*)' for aj, bj in zip(correct[j]['인수수량'], first[j]['인수수량'])])
                        else:
                            acc_cnt2 = '확인필요'
                else :
                    acc_cnt1 = '\n'.join(first[j]['대표수량'])
                    acc_cnt2 = '\n'.join(first[j]['인수수량'])                   
                    
                dif = '' if correct == [] else int(correct[j]['수량'])-int(first[j]['수량'])
                knd_wrt = '' if first[j]['채무증권명칭']=='무보증사채' else first[j]['채무증권명칭']
                rept_no1 = info[i][:14] if info[i][-2:]=='최초' else get_rcept_no(info[i])
                rept_no2 = '' if info[i][-2:]=='최초' else info[i][:14]
                row = {'발행사':company_nm,'회차':seq_no,'등급':crdt_grd,'만기':due,'신고서':rcept_dt, '수요예측(일반공모청약)':'',
                       '납입일':pym_prd,'수량(최초)':face_tot1,'공모희망금리':'','발행조건(확정후)':'','수량(확정)':face_tot2,
                       '발행금리':'','밴드 상단내 수요예측 참여물량':'','수요예측 총참여물량':'','경쟁률':'',
                       '대표주관':acc_nmt1, '대표수량':acc_cnt1,'인수단':acc_nmt2,'인수수량':acc_cnt2,'증액여부':dif,'비고':knd_wrt,
                       '최초보고서':rept_no1,'정정보고서':rept_no2}
                rows.append(row)
            print(info[i])
            st.write('<p style="font-size:14px; color:black">'+'- 문서 '+info[i][:14]+' 추출 완료! ('+info[i][24:-3]+')</p>',unsafe_allow_html=True)
        except Exception as e:
                print(info[i]+'_Error!_'+str(e))
                st.write('<p style="font-size:14px; color:red">'+'- 문서 '+info[i][:14]+'에서 오류 발생! 데이터솔루션부에 문의하세요.</p>',unsafe_allow_html=True)
    df = pd.DataFrame(rows)
    result = df.drop_duplicates(['발행사','회차','최초보고서'],keep='first')
    return(result)

### STEP5. 웹페이지 레이아웃 및 엑셀 형식 설정하기
# 애니메이션 및 보고서 제목 삽입
def load_lottie():
    with open('./resources/report.json', 'r', encoding='utf-8-sig') as st_json:
        return json.load(st_json)

empty1, col1, col2 = st.columns([0.05, 0.3, 0.8])
with empty1:
    st.empty()
with col1:
    lottie = load_lottie()
    st_lottie(lottie, speed=1, loop=True, width=150, height=150, )
with col2:
    ''
    ''
    st.title('공모회사채 발행현황')

# 날짜 선택
start_date = st.date_input('시작일', value=date.today(), max_value = date.today())
max_date = min(start_date+timedelta(days=7), date.today())
end_date = st.date_input('종료일', value=start_date, min_value = start_date, max_value = max_date)

# 조회 및 다운 버튼 생성
if st.button("조회"):
    bgn_de = datetime.strftime(start_date,'%Y%m%d')
    end_de = datetime.strftime(end_date,'%Y%m%d')
    info = get_info(bgn_de, end_de)
    result = get_report(info)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(result, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            if column_cells[0].column_letter == 'E' or column_cells[0].column_letter == 'G' :
                cell.number_format = 'yy/mm/dd'
    ws.column_dimensions['A'].width = ws.column_dimensions['U'].width = ws.column_dimensions['V'].width = ws.column_dimensions['W'].width = 15
    if bgn_de == end_de :
        wb.save('회사채 발행_'+bgn_de+'.xlsx')
        st.dataframe(result)
        with open('회사채 발행_'+bgn_de+'.xlsx', 'rb') as f:
                data = f.read()
                st.download_button(label='다운', data=data, file_name='회사채 발행_'+bgn_de+'.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else :
        wb.save('회사채 발행_'+bgn_de+'_'+end_de+'.xlsx')
        st.dataframe(result)
        with open('회사채 발행_'+bgn_de+'_'+end_de+'.xlsx', 'rb') as f:
                data = f.read()
                st.download_button(label='다운', data=data, file_name='회사채 발행_'+bgn_de+'_'+end_de+'.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
